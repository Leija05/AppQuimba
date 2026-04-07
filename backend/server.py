from fastapi import FastAPI, APIRouter, UploadFile, File, HTTPException, Request
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from starlette.responses import JSONResponse
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
import uuid
from datetime import datetime, timezone
from io import BytesIO
import asyncio
import json
import re
import hmac
import hashlib
import openpyxl
import uvicorn

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')
DATA_DIR = ROOT_DIR / "data"
DATA_DIR.mkdir(parents=True, exist_ok=True)

# Archivos separados para cada sección
LOGISTICA_RECORDS_FILE = DATA_DIR / "logistica_records.json"
TRANSPORTISTA_RECORDS_FILE = DATA_DIR / "transportista_records.json"
LOGISTICA_UPLOADS_FILE = DATA_DIR / "logistica_uploads.json"
TRANSPORTISTA_UPLOADS_FILE = DATA_DIR / "transportista_uploads.json"
CLIENTS_FILE = DATA_DIR / "clients.json"

# Legacy files for migration
RECORDS_FILE = DATA_DIR / "records.json"
UPLOADS_FILE = DATA_DIR / "uploads.json"

_records_lock = asyncio.Lock()

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")


@app.middleware("http")
async def enforce_license(request: Request, call_next):
    if request.url.path in {"/api/license/verify"} or request.url.path.startswith("/docs") or request.url.path.startswith("/openapi"):
        return await call_next(request)
    token = os.environ.get("QUIMBAR_LICENSE_TOKEN", "")
    machine_id = os.environ.get("QUIMBAR_MACHINE_ID", "")
    valid, reason = _validate_token(token, machine_id or None)
    if not valid:
        return JSONResponse(status_code=403, content={"detail": f"Licencia inválida: {reason}"})
    return await call_next(request)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

LICENSE_SECRET = os.environ.get("QUIMBAR_LICENSE_SECRET", "QuimbarToken2026")


# ============== Helper Functions ==============

def _load_json_sync(filepath: Path) -> List[dict]:
    if not filepath.exists():
        return []
    with filepath.open("r", encoding="utf-8") as f:
        data = json.load(f)
    return data if isinstance(data, list) else []


def _save_json_sync(filepath: Path, data: List[dict]) -> None:
    with filepath.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


async def load_json(filepath: Path) -> List[dict]:
    return await asyncio.to_thread(_load_json_sync, filepath)


async def save_json(filepath: Path, data: List[dict]) -> None:
    await asyncio.to_thread(_save_json_sync, filepath, data)


def _normalize_excel_cell(cell_value) -> str:
    if cell_value is None:
        return ""
    return str(cell_value).strip().lower()


def _is_summary_row(row_values) -> bool:
    summary_tokens = (
        "total pendiente",
        "total pagado",
        "total general",
        "tot pendiente",
        "tot pagado",
        "tot general",
    )
    row_text = " ".join(_normalize_excel_cell(value) for value in row_values if value is not None)
    return any(token in row_text for token in summary_tokens) or (
        "total" in row_text and ("pendiente" in row_text or "pagado" in row_text or "general" in row_text)
    )


def _find_header_row(sheet) -> int:
    required_headers = {"fecha", "costo", "transportista", "transporte", "servicio", "status", "estado", "total", "carta", "porte", "shipment"}
    scan_limit = min(sheet.max_row or 0, 25)
    best_row = 1
    best_score = -1

    for row_index in range(1, scan_limit + 1):
        row_values = [_normalize_excel_cell(cell.value) for cell in sheet[row_index]]
        row_text = " ".join(row_values)
        score = sum(1 for value in required_headers if value in row_text)
        if score > best_score:
            best_score = score
            best_row = row_index

    return best_row


def _normalize_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", value.lower())


def _parse_float(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    cleaned = str(value).strip()
    if cleaned in {"", "-"}:
        return 0.0

    cleaned = cleaned.replace("$", "").replace(",", "").strip()
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def _machine_hash(machine_id: str) -> str:
    return hashlib.sha256(machine_id.encode("utf-8")).hexdigest()[:12]


def _validate_token(token: str, machine_id: Optional[str] = None) -> tuple[bool, str]:
    if not token or (not token.startswith("QBM.") and not token.startswith("QBM2.")):
        return False, "Token vacío o formato inválido"
    try:
        parts = token.split(".")
        token_version = parts[0]

        if token_version == "QBM2":
            if len(parts) != 4:
                return False, "Formato QBM2 inválido"
            _, expiry_compact, token_machine_hash, signature = parts
            if machine_id and token_machine_hash != _machine_hash(machine_id):
                return False, "Token no pertenece a esta computadora"
            payload = f"{expiry_compact}.{token_machine_hash}"
        else:
            if len(parts) != 3:
                return False, "Formato QBM inválido"
            _, expiry_compact, signature = parts
            payload = expiry_compact

        expiry_date = datetime.strptime(expiry_compact, "%Y%m%d").date()
        expected = hmac.new(LICENSE_SECRET.encode("utf-8"), payload.encode("utf-8"), hashlib.sha256).hexdigest()[:12]
        if not hmac.compare_digest(expected, signature):
            return False, "Firma inválida"
        if datetime.utcnow().date() > expiry_date:
            return False, "Token caducado"
        return True, "OK"
    except Exception:
        return False, "Token inválido"


# ============== Models for LOGISTICA ==============
class LogisticaRecord(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    fecha: str
    costo: float = 0.0
    carta_porte: str = ""
    servicio: str = ""
    shipment: str = ""
    status: str = "Pendiente"
    total: float = 0.0
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class LogisticaRecordCreate(BaseModel):
    fecha: str
    costo: float = 0.0
    carta_porte: str = ""
    servicio: str = ""
    shipment: str = ""
    status: str = "Pendiente"


class LogisticaRecordUpdate(BaseModel):
    fecha: Optional[str] = None
    costo: Optional[float] = None
    carta_porte: Optional[str] = None
    servicio: Optional[str] = None
    shipment: Optional[str] = None
    status: Optional[str] = None


# ============== Models for TRANSPORTISTA ==============
class TransportistaRecord(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    fecha: str
    costo_t: float = 0.0
    transporte: str = ""
    servicio: str = ""
    costo_l: float = 0.0
    status: str = "Pendiente"
    total: float = 0.0
    saldo_a_favor: float = 0.0
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class TransportistaRecordCreate(BaseModel):
    fecha: str
    costo_t: float = 0.0
    transporte: str = ""
    servicio: str = ""
    costo_l: float = 0.0
    status: str = "Pendiente"
    saldo_a_favor: float = 0.0


class TransportistaRecordUpdate(BaseModel):
    fecha: Optional[str] = None
    costo_t: Optional[float] = None
    transporte: Optional[str] = None
    servicio: Optional[str] = None
    costo_l: Optional[float] = None
    status: Optional[str] = None
    saldo_a_favor: Optional[float] = None


# ============== Response Models ==============
class LogisticaTotalsResponse(BaseModel):
    total_pendiente: float
    total_pagado: float
    total_general: float


class TransportistaTotalsResponse(BaseModel):
    total_pendiente: float
    total_pagado: float
    total_general: float
    total_saldo_a_favor: float


class UploadedFileInfo(BaseModel):
    id: str
    filename: str
    uploaded_at: str
    records_count: int
    section: str = "transportista"


class Client(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nombre: str
    correo: str = ""
    telefono: str = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class ClientCreate(BaseModel):
    nombre: str
    correo: str = ""
    telefono: str = ""


# ============== LOGISTICA Routes ==============

@api_router.get("/logistica/records", response_model=List[LogisticaRecord])
async def get_logistica_records():
    """Get all logistica records"""
    async with _records_lock:
        return await load_json(LOGISTICA_RECORDS_FILE)


@api_router.get("/logistica/records/{record_id}", response_model=LogisticaRecord)
async def get_logistica_record(record_id: str):
    """Get a single logistica record by ID"""
    async with _records_lock:
        records = await load_json(LOGISTICA_RECORDS_FILE)
    record = next((r for r in records if r.get("id") == record_id), None)
    if not record:
        raise HTTPException(status_code=404, detail="Record not found")
    return record


@api_router.post("/logistica/records", response_model=LogisticaRecord)
async def create_logistica_record(data: LogisticaRecordCreate):
    """Create a new logistica record"""
    total = data.costo
    
    record = LogisticaRecord(
        fecha=data.fecha,
        costo=data.costo,
        carta_porte=data.carta_porte,
        servicio=data.servicio,
        shipment=data.shipment,
        status=data.status,
        total=total
    )
    
    doc = record.model_dump()
    async with _records_lock:
        records = await load_json(LOGISTICA_RECORDS_FILE)
        records.append(doc)
        await save_json(LOGISTICA_RECORDS_FILE, records)
    return record


@api_router.put("/logistica/records/{record_id}", response_model=LogisticaRecord)
async def update_logistica_record(record_id: str, data: LogisticaRecordUpdate):
    """Update a logistica record"""
    async with _records_lock:
        records = await load_json(LOGISTICA_RECORDS_FILE)
        record_index = next((i for i, r in enumerate(records) if r.get("id") == record_id), None)

        if record_index is None:
            raise HTTPException(status_code=404, detail="Record not found")

        existing = records[record_index]
        update_data = {k: v for k, v in data.model_dump().items() if v is not None}

        costo = update_data.get("costo", existing.get("costo", 0))
        update_data["total"] = costo

        updated = {**existing, **update_data}
        records[record_index] = updated
        await save_json(LOGISTICA_RECORDS_FILE, records)
        return updated


@api_router.delete("/logistica/records/{record_id}")
async def delete_logistica_record(record_id: str):
    """Delete a logistica record"""
    async with _records_lock:
        records = await load_json(LOGISTICA_RECORDS_FILE)
        filtered = [r for r in records if r.get("id") != record_id]
        if len(filtered) == len(records):
            raise HTTPException(status_code=404, detail="Record not found")
        await save_json(LOGISTICA_RECORDS_FILE, filtered)
    return {"message": "Record deleted successfully"}


@api_router.delete("/logistica/records")
async def delete_all_logistica_records():
    """Delete all logistica records"""
    async with _records_lock:
        records = await load_json(LOGISTICA_RECORDS_FILE)
        deleted_count = len(records)
        await save_json(LOGISTICA_RECORDS_FILE, [])
    return {"message": f"Deleted {deleted_count} records"}


@api_router.get("/logistica/totals", response_model=LogisticaTotalsResponse)
async def get_logistica_totals():
    """Get logistica totals"""
    async with _records_lock:
        records = await load_json(LOGISTICA_RECORDS_FILE)
    
    total_pendiente = sum(r.get("total", 0) for r in records if r.get("status") == "Pendiente")
    total_pagado = sum(r.get("total", 0) for r in records if r.get("status") == "Pagado")
    
    return LogisticaTotalsResponse(
        total_pendiente=total_pendiente,
        total_pagado=total_pagado,
        total_general=total_pendiente + total_pagado
    )


@api_router.post("/logistica/upload-excel")
async def upload_logistica_excel(file: UploadFile = File(...)):
    """Upload Excel file for logistica records"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")
    
    try:
        contents = await file.read()
        workbook = openpyxl.load_workbook(BytesIO(contents), data_only=True)

        column_map = {
            'fecha': ['fecha', 'date'],
            'costo': ['costo', 'cost', 'precio'],
            'carta_porte': ['carta porte', 'cartaporte', 'carta', 'porte', 'cp'],
            'servicio': ['servicio', 'service', 'descripcion'],
            'shipment': ['shipment', 'embarque', 'envio', 'referencia'],
            'status': ['status', 'estado', 'estatus'],
            'total': ['total']
        }
        
        def find_column(field_names, headers):
            normalized_headers = [_normalize_header(h) for h in headers]
            for i, header in enumerate(normalized_headers):
                if not header:
                    continue
                for name in field_names:
                    norm_name = _normalize_header(name)
                    if header == norm_name or header.startswith(norm_name) or norm_name in header:
                        return i
            return -1

        records_imported = 0
        parsed_records = []
        seen_records = set()

        async with _records_lock:
            for sheet in workbook.worksheets:
                header_row = _find_header_row(sheet)
                headers = [_normalize_excel_cell(cell.value) for cell in sheet[header_row]]
                col_indices = {field: find_column(names, headers) for field, names in column_map.items()}

                for row_num, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
                    try:
                        if all(v is None or str(v).strip() == "" for v in row):
                            continue
                        if _is_summary_row(row):
                            continue

                        fecha = row[col_indices['fecha']] if col_indices['fecha'] >= 0 else ""
                        if isinstance(fecha, datetime):
                            fecha = fecha.strftime('%Y-%m-%d')
                        elif fecha:
                            fecha = str(fecha)
                        else:
                            fecha = datetime.now().strftime('%Y-%m-%d')

                        costo = _parse_float(row[col_indices['costo']]) if col_indices['costo'] >= 0 else 0.0
                        carta_porte = str(row[col_indices['carta_porte']] or "") if col_indices['carta_porte'] >= 0 else ""
                        servicio = str(row[col_indices['servicio']] or "") if col_indices['servicio'] >= 0 else ""
                        shipment = str(row[col_indices['shipment']] or "") if col_indices['shipment'] >= 0 else ""
                        status = str(row[col_indices['status']] or "Pendiente")
                        
                        total = costo

                        status = status.strip().capitalize()
                        if status not in ["Pendiente", "Pagado"]:
                            status = "Pendiente"

                        record = LogisticaRecord(
                            fecha=fecha,
                            costo=costo,
                            carta_porte=carta_porte,
                            servicio=servicio,
                            shipment=shipment,
                            status=status,
                            total=total
                        )

                        record_doc = record.model_dump()
                        dedupe_key = (
                            str(record_doc.get("fecha", "")).strip().lower(),
                            str(record_doc.get("carta_porte", "")).strip().lower(),
                            str(record_doc.get("servicio", "")).strip().lower(),
                            round(float(record_doc.get("costo", 0) or 0), 2),
                        )

                        if dedupe_key not in seen_records and (servicio.strip() or carta_porte.strip() or costo != 0):
                            seen_records.add(dedupe_key)
                            parsed_records.append(record_doc)
                            records_imported += 1
                    except Exception as e:
                        logger.error(f"Error parsing row {row_num}: {e}")

            await save_json(LOGISTICA_RECORDS_FILE, parsed_records)

            uploads = await load_json(LOGISTICA_UPLOADS_FILE)
            uploads.append({
                "id": str(uuid.uuid4()),
                "filename": file.filename,
                "uploaded_at": datetime.now(timezone.utc).isoformat(),
                "records": parsed_records,
                "section": "logistica"
            })
            await save_json(LOGISTICA_UPLOADS_FILE, uploads)

        return {
            "message": f"Imported {records_imported} logistica records",
            "records_imported": records_imported,
            "section": "logistica"
        }
        
    except Exception as e:
        logger.error(f"Error processing Excel file: {e}")
        raise HTTPException(status_code=400, detail=f"Error processing file: {str(e)}")


@api_router.get("/logistica/uploads", response_model=List[UploadedFileInfo])
async def get_logistica_uploads():
    async with _records_lock:
        uploads = await load_json(LOGISTICA_UPLOADS_FILE)
    return [
        UploadedFileInfo(
            id=u.get("id"),
            filename=u.get("filename", "Sin nombre"),
            uploaded_at=u.get("uploaded_at", ""),
            records_count=len(u.get("records", [])),
            section="logistica"
        )
        for u in sorted(uploads, key=lambda x: x.get("uploaded_at", ""), reverse=True)
    ]


@api_router.post("/logistica/uploads/{upload_id}/load")
async def load_logistica_upload(upload_id: str):
    async with _records_lock:
        uploads = await load_json(LOGISTICA_UPLOADS_FILE)
        upload = next((u for u in uploads if u.get("id") == upload_id), None)
        if not upload:
            raise HTTPException(status_code=404, detail="Upload not found")
        await save_json(LOGISTICA_RECORDS_FILE, upload.get("records", []))
    return {"message": "Archivo cargado", "records_loaded": len(upload.get("records", []))}


@api_router.delete("/logistica/uploads/{upload_id}")
async def delete_logistica_upload(upload_id: str):
    async with _records_lock:
        uploads = await load_json(LOGISTICA_UPLOADS_FILE)
        filtered = [u for u in uploads if u.get("id") != upload_id]
        if len(filtered) == len(uploads):
            raise HTTPException(status_code=404, detail="Upload not found")
        await save_json(LOGISTICA_UPLOADS_FILE, filtered)
    return {"message": "Archivo eliminado"}


@api_router.delete("/logistica/uploads")
async def delete_all_logistica_uploads():
    async with _records_lock:
        uploads = await load_json(LOGISTICA_UPLOADS_FILE)
        deleted_count = len(uploads)
        await save_json(LOGISTICA_UPLOADS_FILE, [])
    return {"message": f"Deleted {deleted_count} uploads"}


# ============== TRANSPORTISTA Routes ==============

@api_router.get("/transportista/records", response_model=List[TransportistaRecord])
async def get_transportista_records():
    """Get all transportista records"""
    async with _records_lock:
        return await load_json(TRANSPORTISTA_RECORDS_FILE)


@api_router.get("/transportista/records/{record_id}", response_model=TransportistaRecord)
async def get_transportista_record(record_id: str):
    """Get a single transportista record by ID"""
    async with _records_lock:
        records = await load_json(TRANSPORTISTA_RECORDS_FILE)
    record = next((r for r in records if r.get("id") == record_id), None)
    if not record:
        raise HTTPException(status_code=404, detail="Record not found")
    return record


@api_router.post("/transportista/records", response_model=TransportistaRecord)
async def create_transportista_record(data: TransportistaRecordCreate):
    """Create a new transportista record"""
    total = data.costo_l
    saldo_a_favor = data.costo_l - data.costo_t
    
    record = TransportistaRecord(
        fecha=data.fecha,
        costo_t=data.costo_t,
        transporte=data.transporte,
        servicio=data.servicio,
        costo_l=data.costo_l,
        status=data.status,
        total=total,
        saldo_a_favor=saldo_a_favor
    )
    
    doc = record.model_dump()
    async with _records_lock:
        records = await load_json(TRANSPORTISTA_RECORDS_FILE)
        records.append(doc)
        await save_json(TRANSPORTISTA_RECORDS_FILE, records)
    return record


@api_router.put("/transportista/records/{record_id}", response_model=TransportistaRecord)
async def update_transportista_record(record_id: str, data: TransportistaRecordUpdate):
    """Update a transportista record"""
    async with _records_lock:
        records = await load_json(TRANSPORTISTA_RECORDS_FILE)
        record_index = next((i for i, r in enumerate(records) if r.get("id") == record_id), None)

        if record_index is None:
            raise HTTPException(status_code=404, detail="Record not found")

        existing = records[record_index]
        update_data = {k: v for k, v in data.model_dump().items() if v is not None}

        costo_t = update_data.get("costo_t", existing.get("costo_t", 0))
        costo_l = update_data.get("costo_l", existing.get("costo_l", 0))
        
        update_data["total"] = costo_l
        update_data["saldo_a_favor"] = costo_l - costo_t

        updated = {**existing, **update_data}
        records[record_index] = updated
        await save_json(TRANSPORTISTA_RECORDS_FILE, records)
        return updated


@api_router.delete("/transportista/records/{record_id}")
async def delete_transportista_record(record_id: str):
    """Delete a transportista record"""
    async with _records_lock:
        records = await load_json(TRANSPORTISTA_RECORDS_FILE)
        filtered = [r for r in records if r.get("id") != record_id]
        if len(filtered) == len(records):
            raise HTTPException(status_code=404, detail="Record not found")
        await save_json(TRANSPORTISTA_RECORDS_FILE, filtered)
    return {"message": "Record deleted successfully"}


@api_router.delete("/transportista/records")
async def delete_all_transportista_records():
    """Delete all transportista records"""
    async with _records_lock:
        records = await load_json(TRANSPORTISTA_RECORDS_FILE)
        deleted_count = len(records)
        await save_json(TRANSPORTISTA_RECORDS_FILE, [])
    return {"message": f"Deleted {deleted_count} records"}


@api_router.get("/transportista/totals", response_model=TransportistaTotalsResponse)
async def get_transportista_totals():
    """Get transportista totals"""
    async with _records_lock:
        records = await load_json(TRANSPORTISTA_RECORDS_FILE)
    
    total_pendiente = sum(r.get("total", 0) for r in records if r.get("status") == "Pendiente")
    total_pagado = sum(r.get("total", 0) for r in records if r.get("status") == "Pagado")
    total_saldo_a_favor = sum(r.get("saldo_a_favor", 0) for r in records)
    
    return TransportistaTotalsResponse(
        total_pendiente=total_pendiente,
        total_pagado=total_pagado,
        total_general=total_pendiente + total_pagado,
        total_saldo_a_favor=total_saldo_a_favor
    )


@api_router.post("/transportista/upload-excel")
async def upload_transportista_excel(file: UploadFile = File(...)):
    """Upload Excel file for transportista records"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")
    
    try:
        contents = await file.read()
        workbook = openpyxl.load_workbook(BytesIO(contents), data_only=True)

        column_map = {
            'fecha': ['fecha', 'date'],
            'costo_t': ['costo t', 'costot', 'costo transporte', 'costo'],
            'transporte': ['transporte', 'transportista', 'carrier'],
            'servicio': ['servicio', 'service', 'descripcion'],
            'costo_l': ['costo l', 'costol', 'costo local', 'costo logistica'],
            'status': ['status', 'estado', 'estatus'],
            'total': ['total'],
            'saldo_a_favor': ['saldo a favor', 'saldo', 'balance']
        }
        
        def find_column(field_names, headers):
            normalized_headers = [_normalize_header(h) for h in headers]
            for i, header in enumerate(normalized_headers):
                if not header:
                    continue
                for name in field_names:
                    norm_name = _normalize_header(name)
                    if header == norm_name or header.startswith(norm_name) or norm_name in header:
                        return i
            return -1

        records_imported = 0
        parsed_records = []
        seen_records = set()

        async with _records_lock:
            for sheet in workbook.worksheets:
                header_row = _find_header_row(sheet)
                headers = [_normalize_excel_cell(cell.value) for cell in sheet[header_row]]
                col_indices = {field: find_column(names, headers) for field, names in column_map.items()}
                
                # Handle case where costo_t and costo_l might be the same column
                normalized_headers = [_normalize_header(h) for h in headers]
                costo_candidates = [i for i, h in enumerate(normalized_headers) if h.startswith("costo")]
                
                if col_indices["costo_t"] == col_indices["costo_l"] and len(costo_candidates) >= 2:
                    col_indices["costo_t"] = costo_candidates[0]
                    col_indices["costo_l"] = costo_candidates[1]
                elif col_indices["costo_t"] == col_indices["costo_l"]:
                    col_indices["costo_l"] = -1

                if col_indices["costo_t"] < 0 and costo_candidates:
                    col_indices["costo_t"] = costo_candidates[0]

                if col_indices["costo_l"] < 0 and len(costo_candidates) >= 2:
                    col_indices["costo_l"] = costo_candidates[1]

                for row_num, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
                    try:
                        if all(v is None or str(v).strip() == "" for v in row):
                            continue
                        if _is_summary_row(row):
                            continue

                        fecha = row[col_indices['fecha']] if col_indices['fecha'] >= 0 else ""
                        if isinstance(fecha, datetime):
                            fecha = fecha.strftime('%Y-%m-%d')
                        elif fecha:
                            fecha = str(fecha)
                        else:
                            fecha = datetime.now().strftime('%Y-%m-%d')

                        costo_t = _parse_float(row[col_indices['costo_t']]) if col_indices['costo_t'] >= 0 else 0.0
                        costo_l = _parse_float(row[col_indices['costo_l']]) if col_indices['costo_l'] >= 0 else 0.0
                        transporte = str(row[col_indices['transporte']] or "") if col_indices['transporte'] >= 0 else ""
                        servicio = str(row[col_indices['servicio']] or "") if col_indices['servicio'] >= 0 else ""
                        status = str(row[col_indices['status']] or "Pendiente")
                        
                        total = costo_l
                        saldo_a_favor = costo_l - costo_t

                        status = status.strip().capitalize()
                        if status not in ["Pendiente", "Pagado"]:
                            status = "Pendiente"

                        record = TransportistaRecord(
                            fecha=fecha,
                            costo_t=costo_t,
                            transporte=transporte,
                            servicio=servicio,
                            costo_l=costo_l,
                            status=status,
                            total=total,
                            saldo_a_favor=saldo_a_favor
                        )

                        record_doc = record.model_dump()
                        dedupe_key = (
                            str(record_doc.get("fecha", "")).strip().lower(),
                            str(record_doc.get("transporte", "")).strip().lower(),
                            str(record_doc.get("servicio", "")).strip().lower(),
                            round(float(record_doc.get("costo_t", 0) or 0), 2),
                            round(float(record_doc.get("costo_l", 0) or 0), 2),
                        )

                        if dedupe_key not in seen_records and (servicio.strip() or transporte.strip() or costo_t != 0 or costo_l != 0):
                            seen_records.add(dedupe_key)
                            parsed_records.append(record_doc)
                            records_imported += 1
                    except Exception as e:
                        logger.error(f"Error parsing row {row_num}: {e}")

            await save_json(TRANSPORTISTA_RECORDS_FILE, parsed_records)

            uploads = await load_json(TRANSPORTISTA_UPLOADS_FILE)
            uploads.append({
                "id": str(uuid.uuid4()),
                "filename": file.filename,
                "uploaded_at": datetime.now(timezone.utc).isoformat(),
                "records": parsed_records,
                "section": "transportista"
            })
            await save_json(TRANSPORTISTA_UPLOADS_FILE, uploads)

        return {
            "message": f"Imported {records_imported} transportista records",
            "records_imported": records_imported,
            "section": "transportista"
        }
        
    except Exception as e:
        logger.error(f"Error processing Excel file: {e}")
        raise HTTPException(status_code=400, detail=f"Error processing file: {str(e)}")


@api_router.get("/transportista/uploads", response_model=List[UploadedFileInfo])
async def get_transportista_uploads():
    async with _records_lock:
        uploads = await load_json(TRANSPORTISTA_UPLOADS_FILE)
    return [
        UploadedFileInfo(
            id=u.get("id"),
            filename=u.get("filename", "Sin nombre"),
            uploaded_at=u.get("uploaded_at", ""),
            records_count=len(u.get("records", [])),
            section="transportista"
        )
        for u in sorted(uploads, key=lambda x: x.get("uploaded_at", ""), reverse=True)
    ]


@api_router.post("/transportista/uploads/{upload_id}/load")
async def load_transportista_upload(upload_id: str):
    async with _records_lock:
        uploads = await load_json(TRANSPORTISTA_UPLOADS_FILE)
        upload = next((u for u in uploads if u.get("id") == upload_id), None)
        if not upload:
            raise HTTPException(status_code=404, detail="Upload not found")
        await save_json(TRANSPORTISTA_RECORDS_FILE, upload.get("records", []))
    return {"message": "Archivo cargado", "records_loaded": len(upload.get("records", []))}


@api_router.delete("/transportista/uploads/{upload_id}")
async def delete_transportista_upload(upload_id: str):
    async with _records_lock:
        uploads = await load_json(TRANSPORTISTA_UPLOADS_FILE)
        filtered = [u for u in uploads if u.get("id") != upload_id]
        if len(filtered) == len(uploads):
            raise HTTPException(status_code=404, detail="Upload not found")
        await save_json(TRANSPORTISTA_UPLOADS_FILE, filtered)
    return {"message": "Archivo eliminado"}


@api_router.delete("/transportista/uploads")
async def delete_all_transportista_uploads():
    async with _records_lock:
        uploads = await load_json(TRANSPORTISTA_UPLOADS_FILE)
        deleted_count = len(uploads)
        await save_json(TRANSPORTISTA_UPLOADS_FILE, [])
    return {"message": f"Deleted {deleted_count} uploads"}


@api_router.get("/clients", response_model=List[Client])
async def get_clients():
    async with _records_lock:
        return await load_json(CLIENTS_FILE)


@api_router.post("/clients", response_model=Client)
async def create_client(data: ClientCreate):
    client = Client(nombre=data.nombre, correo=data.correo, telefono=data.telefono)
    async with _records_lock:
        clients = await load_json(CLIENTS_FILE)
        clients.append(client.model_dump())
        await save_json(CLIENTS_FILE, clients)
    return client


@api_router.get("/license/verify")
async def verify_license():
    token = os.environ.get("QUIMBAR_LICENSE_TOKEN", "")
    machine_id = os.environ.get("QUIMBAR_MACHINE_ID", "")
    valid, reason = _validate_token(token, machine_id or None)
    return {"valid": valid, "reason": reason}


# ============== Legacy/Common Routes ==============

@api_router.get("/")
async def root():
    return {"message": "Sistema de Quimbar API - Logística y Transportista"}


# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("shutdown")
async def shutdown_db_client():
    return None


if __name__ == "__main__":
    host = os.environ.get("HOST", "127.0.0.1")
    port = int(os.environ.get("PORT", "8000"))
    uvicorn.run(app, host=host, port=port)
